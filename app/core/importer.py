"""Excel/CSV 导入解析（P1.4）。

把上传文件内容解析为行 dict 列表（键为中文列头）。列头为模板约定，
与 scripts/make_import_templates.py 生成的模板一致，改动需同步。

解析层只做结构/基本字段检查，业务校验（受控值、去重）在 service 层逐行收集：
- ImportFormatError → 422（非 Excel / 缺列 / 编码问题）
- ImportEmptyError → 400（文件无有效数据行）
"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

import openpyxl

# ==================== 导入模板列头（唯一约定，与下载模板一致） ====================
# 对外模板不含"编号"列（避免用户困惑）；service 解析后统一生成随机 ID。
# 兼容：若上传文件仍含"编号"列（内部验收模板/合成数据复用 ID），_norm_rows 保留该键，
# service 优先复用（row.get("编号")），因此多余列无害。
EXPERT_EXCEL_HEADERS = ["姓名", "单位", "地区", "从业年限", "专业标签", "身份证号", "邮箱", "电话"]
SUPPLIER_EXCEL_HEADERS = ["企业名称", "统一社会信用代码", "法定代表人", "所属行业", "企业规模"]
CONFLICT_CSV_HEADERS = ["姓名", "企业名称", "统一社会信用代码", "关系类型", "职位", "持股比例"]


class ImportFormatError(ValueError):
    """文件格式错误（非 Excel / 缺列 / 编码）→ 422。"""


class ImportEmptyError(ValueError):
    """文件无有效数据行 → 400。"""


def _check_headers(headers: list[str], expected: list[str], label: str) -> None:
    """校验列头完整。缺列即格式错误（模板约定，缺列无法安全解析）。"""
    missing = [h for h in expected if h not in headers]
    if missing:
        raise ImportFormatError(f"{label}文件缺少列: {missing}，期望列头 {expected}")


def _norm_rows(headers: list[str], rows: list[tuple]) -> list[dict]:
    """行 tuple → dict，空行跳过。None 规范为空字符串。"""
    out: list[dict] = []
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({k: ("" if v is None else str(v).strip()) for k, v in zip(headers, row)})
    return out


def parse_expert_excel(content: bytes) -> list[dict]:
    """解析专家导入 Excel → 行 dict 列表。"""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001  非 xlsx/损坏文件统一报格式错误
        raise ImportFormatError(f"无法解析 Excel 文件: {e}")
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows, ())]
    _check_headers(headers, EXPERT_EXCEL_HEADERS, "专家导入")
    out = _norm_rows(headers, list(rows))
    if not out:
        raise ImportEmptyError("Excel 无有效数据行")
    return out


def parse_supplier_excel(content: bytes) -> list[dict]:
    """解析供应商导入 Excel → 行 dict 列表。"""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ImportFormatError(f"无法解析 Excel 文件: {e}")
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows, ())]
    _check_headers(headers, SUPPLIER_EXCEL_HEADERS, "供应商导入")
    out = _norm_rows(headers, list(rows))
    if not out:
        raise ImportEmptyError("Excel 无有效数据行")
    return out


def parse_conflict_csv(content: bytes) -> list[dict]:
    """解析企查查风格 CSV → 行 dict 列表。utf-8-sig 兼容 BOM。"""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ImportFormatError("CSV 必须为 UTF-8 编码（可含 BOM）")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ImportEmptyError("CSV 无表头")
    headers = [h.strip() for h in reader.fieldnames]
    _check_headers(headers, CONFLICT_CSV_HEADERS, "冲突")
    out = []
    for row in reader:
        cleaned = {k.strip(): (v or "").strip() for k, v in row.items()}
        if all(not v for v in cleaned.values()):
            continue
        out.append(cleaned)
    if not out:
        raise ImportEmptyError("CSV 无有效数据行")
    return out
